"""
Excel parser for Epicenter Valuation Dashboard.

Parses Bloomberg BQL-exported Excel workbooks into the dashboard JSON format.
Handles the transposed layout where tickers are in columns and dates are in rows.

The Excel workbook may contain any combination of the following sheets:

    Data sheets (tickers in column headers, dates in column B):
        EV - Rev   -> er (EV/Revenue ratio)
        EV - GP    -> eg (EV/Gross Profit ratio)
        PE         -> pe (Price/EPS ratio)
        Rev Growth -> rg (Revenue growth, decimal)
        EPS Growth -> xg (EPS growth, decimal)
        Forward EPS -> fe (Forward EPS, absolute $)

    Lookup sheet:
        Industries -> two-column ticker-to-industry mapping

Sheet layout for data sheets:
    Rows 1-5(ish): metadata (start date, end date, periods, etc.)
    Row 6: metric label row (ignored)
    Row 7: ticker names (e.g. "ADBE US Equity") in columns C, D, E, ...
    Row 8: either a duplicate ticker row OR the first date row
    Row 8/9+: dates in column B, data values in columns C+ for each ticker

Auto-detection: if cell(row=8, col=2) is a datetime, data starts at row 8.
Otherwise, data starts at row 9 (row 8 is a duplicate ticker header).
"""

from __future__ import annotations

import io
import logging
import math
from datetime import datetime
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Mapping from sheet name to metric key in the output JSON.
SHEET_METRIC_MAP: dict[str, str] = {
    "EV - Rev": "er",
    "EV - GP": "eg",
    "PE": "pe",
    "Rev Growth": "rg",
    "EPS Growth": "xg",
    "Forward EPS": "fe",
}

# All metric keys that each ticker must have in the output.
ALL_METRIC_KEYS = ("er", "eg", "pe", "rg", "xg", "fe")

# Strings that should be treated as null/missing values.
_NULL_STRINGS = {
    "#N/A",
    "#VALUE!",
    "#REF!",
    "#DIV/0!",
    "#NULL!",
    "#NAME?",
    "#NUM!",
    "N/A",
    "",
}


def _is_datetime(val: Any) -> bool:
    """Check if a value is a datetime object."""
    return isinstance(val, datetime)


def _clean_ticker(raw: Any) -> str | None:
    """
    Clean a raw ticker value from the spreadsheet.

    Strips the " US Equity" suffix and whitespace.
    Returns None if the value is not a valid ticker string.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in _NULL_STRINGS:
        return None
    # Strip Bloomberg suffix
    if s.endswith(" US Equity"):
        s = s[: -len(" US Equity")].strip()
    # Reject very short or suspicious values (like 'x', single chars that aren't real tickers)
    if len(s) < 1:
        return None
    return s


def _clean_value(val: Any) -> float | None:
    """
    Convert a cell value to a float or None.

    Handles #N/A strings, NaN, None, empty strings, etc.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return None
        return round(float(val), 4)
    s = str(val).strip()
    if s in _NULL_STRINGS:
        return None
    try:
        f = float(s)
        if math.isnan(f) or math.isinf(f):
            return None
        return round(f, 4)
    except (ValueError, TypeError):
        return None


def _format_date(dt: datetime) -> str:
    """Format a datetime as YYYY-MM-DD string."""
    return dt.strftime("%Y-%m-%d")


def _parse_data_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[list[str], list[str], dict[str, list[float | None]]]:
    """
    Parse a data sheet with tickers in columns and dates in rows.

    Returns:
        (dates, tickers, data_by_ticker)
        - dates: list of "YYYY-MM-DD" strings
        - tickers: list of cleaned ticker strings (preserving column order)
        - data_by_ticker: {ticker: [value_per_date]}
    """
    max_col = ws.max_column
    max_row = ws.max_row

    if max_col is None or max_row is None or max_col < 3 or max_row < 9:
        logger.warning("Sheet '%s' is too small to contain data", ws.title)
        return [], [], {}

    # --- Step 1: Find tickers in row 7, columns C onward ---
    raw_tickers: list[tuple[int, str]] = []  # (column_index, cleaned_ticker)
    for col in range(3, max_col + 1):
        raw = ws.cell(row=7, column=col).value
        ticker = _clean_ticker(raw)
        if ticker is not None:
            raw_tickers.append((col, ticker))

    if not raw_tickers:
        # Fallback: try row 8
        for col in range(3, max_col + 1):
            raw = ws.cell(row=8, column=col).value
            ticker = _clean_ticker(raw)
            if ticker is not None:
                raw_tickers.append((col, ticker))

    if not raw_tickers:
        logger.warning("Sheet '%s': no tickers found", ws.title)
        return [], [], {}

    # --- Step 2: Determine data start row ---
    # If cell(8, 2) is a datetime, data starts at row 8.
    # Otherwise try row 9.
    data_start_row = None
    for candidate in [8, 9, 10]:
        val = ws.cell(row=candidate, column=2).value
        if _is_datetime(val):
            data_start_row = candidate
            break

    if data_start_row is None:
        logger.warning("Sheet '%s': cannot find date column start", ws.title)
        return [], [], {}

    # --- Step 3: Read dates from column B ---
    dates: list[str] = []
    date_rows: list[int] = []
    for row in range(data_start_row, max_row + 1):
        val = ws.cell(row=row, column=2).value
        if val is None:
            continue
        if _is_datetime(val):
            dates.append(_format_date(val))
            date_rows.append(row)
        else:
            # Try parsing a string date
            s = str(val).strip()
            if s:
                try:
                    dt = datetime.strptime(s[:10], "%Y-%m-%d")
                    dates.append(_format_date(dt))
                    date_rows.append(row)
                except ValueError:
                    # Not a date, stop reading
                    break

    if not dates:
        logger.warning("Sheet '%s': no dates found", ws.title)
        return [], [], {}

    # --- Step 4: Read data values ---
    tickers: list[str] = []
    data_by_ticker: dict[str, list[float | None]] = {}

    for col_idx, ticker in raw_tickers:
        if ticker in data_by_ticker:
            # Duplicate ticker column -- skip the later one
            continue
        values: list[float | None] = []
        for row in date_rows:
            cell_val = ws.cell(row=row, column=col_idx).value
            values.append(_clean_value(cell_val))
        tickers.append(ticker)
        data_by_ticker[ticker] = values

    return dates, tickers, data_by_ticker


def _parse_industries_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> dict[str, str]:
    """
    Parse the Industries lookup sheet.

    Expected layout: two columns -- Ticker (col A) and Industry (col B).
    First row is a header. Tickers may have " US Equity" suffix.
    """
    industries: dict[str, str] = {}
    max_row = ws.max_row or 1

    for row in range(2, max_row + 1):  # Skip header row
        raw_ticker = ws.cell(row=row, column=1).value
        raw_industry = ws.cell(row=row, column=2).value
        ticker = _clean_ticker(raw_ticker)
        if ticker is None:
            continue
        if raw_industry is None:
            continue
        industry = str(raw_industry).strip()
        if industry and industry not in _NULL_STRINGS:
            industries[ticker] = industry

    return industries


def parse_excel(file_content: bytes) -> dict:
    """
    Parse an Excel workbook (as bytes) into the dashboard JSON dict.

    The output schema:
    {
        "dates": ["2015-01-01", ...],
        "tickers": ["AAPL", ...],
        "industries": {"AAPL": "Technology", ...},
        "fm": {
            "AAPL": {
                "er": [12.3, null, ...],
                "eg": [25.1, null, ...],
                "pe": [18.5, null, ...],
                "rg": [0.05, null, ...],
                "xg": [0.12, null, ...],
                "fe": [6.50, null, ...]
            }
        }
    }
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(file_content),
        data_only=True,
        read_only=False,
    )

    sheet_names = set(wb.sheetnames)
    logger.info("Workbook sheets: %s", wb.sheetnames)

    # --- Parse each data sheet ---
    # Accumulate: dates (union), tickers (union), per-ticker per-metric data
    all_dates_set: set[str] = set()
    sheet_results: dict[
        str, tuple[list[str], list[str], dict[str, list[float | None]]]
    ] = {}

    for sheet_name, metric_key in SHEET_METRIC_MAP.items():
        if sheet_name not in sheet_names:
            logger.info(
                "Sheet '%s' not found, skipping metric '%s'", sheet_name, metric_key
            )
            sheet_results[metric_key] = ([], [], {})
            continue

        ws = wb[sheet_name]
        dates, tickers, data = _parse_data_sheet(ws)
        sheet_results[metric_key] = (dates, tickers, data)
        all_dates_set.update(dates)
        logger.info(
            "Sheet '%s' -> metric '%s': %d dates, %d tickers",
            sheet_name,
            metric_key,
            len(dates),
            len(tickers),
        )

    # --- Parse Industries sheet ---
    industries: dict[str, str] = {}
    if "Industries" in sheet_names:
        industries = _parse_industries_sheet(wb["Industries"])
        logger.info("Industries: %d mappings", len(industries))

    wb.close()

    # --- Build unified date list (sorted chronologically) ---
    all_dates = sorted(all_dates_set)
    num_dates = len(all_dates)
    date_index_map = {d: i for i, d in enumerate(all_dates)}

    # --- Build unified ticker list (union across all sheets) ---
    all_tickers_set: set[str] = set()
    for metric_key, (_, tickers, _) in sheet_results.items():
        all_tickers_set.update(tickers)
    # Also include tickers from the industries sheet
    all_tickers_set.update(industries.keys())
    # Sort alphabetically for deterministic output
    all_tickers = sorted(all_tickers_set)

    # --- Build fm (financial metrics) dict ---
    fm: dict[str, dict[str, list[float | None]]] = {}

    for ticker in all_tickers:
        fm[ticker] = {}
        for metric_key in ALL_METRIC_KEYS:
            sheet_dates, sheet_tickers, sheet_data = sheet_results.get(
                metric_key, ([], [], {})
            )

            if ticker not in sheet_data:
                # Ticker not present in this sheet -- fill with nulls
                fm[ticker][metric_key] = [None] * num_dates
                continue

            # Map this sheet's data into the unified date array
            values = [None] * num_dates
            ticker_values = sheet_data[ticker]
            for i, date_str in enumerate(sheet_dates):
                if i < len(ticker_values):
                    unified_idx = date_index_map.get(date_str)
                    if unified_idx is not None:
                        values[unified_idx] = ticker_values[i]
            fm[ticker][metric_key] = values

    # --- Strip tickers where every metric is entirely null ---
    empty_tickers = [
        t
        for t in all_tickers
        if all(v is None for key in ALL_METRIC_KEYS for v in fm[t][key])
    ]
    for t in empty_tickers:
        del fm[t]
    all_tickers = [t for t in all_tickers if t not in set(empty_tickers)]

    if empty_tickers:
        logger.info("Stripped %d all-null tickers", len(empty_tickers))

    # --- Build output ---
    result = {
        "dates": all_dates,
        "tickers": all_tickers,
        "industries": industries,
        "fm": fm,
    }

    logger.info(
        "Parsed: %d dates, %d tickers, %d industries, %d metrics",
        len(all_dates),
        len(all_tickers),
        len(industries),
        len(ALL_METRIC_KEYS),
    )

    return result
